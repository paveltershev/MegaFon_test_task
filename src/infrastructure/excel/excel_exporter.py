from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from src.domain.interfaces import Exporter
from src.domain.models import WeatherReport

class ExcelExporter(Exporter):
    def export(self, report: WeatherReport) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Прогноз погоды"

        ws.merge_cells("A1:H1")
        ws["A1"] = f"Прогноз погоды для города {report.city}"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Дата", "Магн. поле", "Утро", "День", "Вечер", "Ночь", "Ср. t днём", "Изменение давления"]
        ws.append([""] + headers)

        for day in report.days:
            row = [
                day.date,
                str(day.magnetic_field),
                f"{day.periods[1].temperature}°C, {day.periods[1].pressure_mm} мм, {day.periods[1].humidity_percent}%, {day.periods[1].condition}",
                f"{day.periods[2].temperature}°C, {day.periods[2].pressure_mm} мм, {day.periods[2].humidity_percent}%, {day.periods[2].condition}",
                f"{day.periods[3].temperature}°C, {day.periods[3].pressure_mm} мм, {day.periods[3].humidity_percent}%, {day.periods[3].condition}",
                f"{day.periods[0].temperature}°C, {day.periods[0].pressure_mm} мм, {day.periods[0].humidity_percent}%, {day.periods[0].condition}",
                str(day.avg_day_temp),
                day.pressure_change_alert
            ]
            ws.append([""] + row)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        filename = f"weather_{report.city.lower()}.xlsx"
        wb.save(filename)
        return filename